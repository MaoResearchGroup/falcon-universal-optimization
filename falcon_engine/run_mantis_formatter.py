import shutil
import openpyxl
import os
import platform 
import subprocess 
from falcon_engine.utilities import print_slowly
from openpyxl.utils import get_column_letter
import pandas as pd

def run_mantis_formatter_pipeline(RUN_NAME, input_param_names, optimized_formulations): 
    '''
    Function to format optimized formulations for MANTIS liquid handler input. 
    Completes first step of copying the template Excel file, then fills in the formulation parameters.
    Following specification of stock concentrations, converts formulation template into mantis-ready csv. 
    '''
    print_slowly("\n\n--- FORMATTING OPTIMIZED FORMULATIONS FOR MANTIS ---")

    # Define source and destination file paths
    src = os.path.join("exp_templates", "template_formulation_sheet.xlsx")
    dst = os.path.join("exp_templates", f"{RUN_NAME}_formulation_sheet.xlsx")
    
    # Copy the file
    shutil.copy(src, dst)
    print(f"Formulation template copied as {dst}")

    excel_param_order = ['NP_ratio', 'IL+HL', 'HL_IL+HL', 'PEG_PEG+Chol']
    # Create a lookup: param name -> index in optimized_formulations
    param_index_lookup = {name: idx for idx, name in enumerate(input_param_names)}

    wb = openpyxl.load_workbook(dst)
    ws = wb["Formulations"] 
    reversed_x = optimized_formulations[input_param_names].values.tolist()
    opt_methods = optimized_formulations["opt_method"].tolist()

    #take user input for IL and HL names
    IL_name = input("Enter the name of your Ionizable Lipid used (IL)\n(Choose SM102, Dlin, or ALC0315): ")
    HL_name = input(
        "Enter the name of your Helper Lipid used (HL)\n"
        "(Choose from: DOTAP, DSPC, 18PG, DOPE, DDAB, 14PA, 18MP): "
    )

    for i in range(len(reversed_x)): 
      for j, param_name in enumerate(excel_param_order):
        row_idx = 10 + j
        col_letter = get_column_letter(3 + i)
        value = reversed_x[i][param_index_lookup[param_name]]
        ws[f"{col_letter}{row_idx}"] = round(value, 3)
        ws[f"{col_letter}6"] = IL_name 
        ws[f"{col_letter}7"] = HL_name 
        ws[f"{col_letter}8"] = "Chol"
        ws[f"{col_letter}9"] = "DMG_PEG" 

        ws[f"{col_letter}3"] = i 
        ws[f"{col_letter}4"] =  opt_methods[i]

    wb.save(dst)
    print(f"Optimized formulations formatted for MANTIS and saved to {dst}")

    open_excel_file(dst)
    input("\nPlease edit the Excel file to fill in stock conentrations, etc.\nOnce you're done, save and close it, then press Enter to continue...")
    
    # format the file to mantis_ready_csv
    # load the file again 
    wb = openpyxl.load_workbook(dst, data_only = True)
    ws = wb["Formulations"] 

    param_rows = {
    "Ionizable_lipid": 54,
    "Helper_lipid": 55,
    "Chol": 56,
    "PEG": 57,
    }
    volume_rows = {
        "Ionizable_lipid": 63,
        "Helper_lipid": 64,
        "Chol": 65,
        "PEG": 66,
        "EtOH": 67,
    }

    # === DETERMINE NUMBER OF FORMULATIONS ===
    # Start from column C (index 3), go right until empty
    formulation_cols = []
    col = 3
    while ws.cell(row=54, column=col).value is not None:
        formulation_cols.append(col)
        col += 1

    # === BUILD UNIQUE COLUMN HEADERS ===
    unique_headers = []

    # 1. Get all unique stock concentrations for naming
    stock_map = {}
    for name, row in param_rows.items():
        stock_map[name] = {}
        for col in formulation_cols:
            conc = ws.cell(row=row, column=col).value
            if name == "Ionizable_lipid":
                stock_map[name][conc] = f"{IL_name}_{conc}mg_ml"
            elif name == "Helper_lipid":
                stock_map[name][conc] = f"{HL_name}_{conc}mg_ml"
            else:
                stock_map[name][conc] = f"{name}_{conc}mg_ml"

    # 2. Flatten and deduplicate column names (preserve order)
    for name in ["Ionizable_lipid", "Helper_lipid", "Chol", "PEG"]:
        for conc, label in stock_map[name].items():
            if label not in unique_headers:
                unique_headers.append(label)

    # Always include ethanol
    unique_headers.append("EtOH_HV")
    unique_headers.append("EtOH_LV")

    # === BUILD DATAFRAME ===
    rows = []
    for f_idx, col in enumerate(formulation_cols):
        row_dict = {}

        # Get volumes by stock conc
        for name in ["Ionizable_lipid", "Helper_lipid", "Chol", "PEG"]:
            stock = ws.cell(row=param_rows[name], column=col).value
            raw_val = ws.cell(row=volume_rows[name], column=col).value
            try:
                vol = float(raw_val)
            except (TypeError, ValueError):
                vol = 0.0
            if name == "Ionizable_lipid":
                label = f"{IL_name}_{stock}mg_ml"
            elif name == "Helper_lipid":
                label = f"{HL_name}_{stock}mg_ml"
            else:
                label = f"{name}_{stock}mg_ml"
            row_dict[label] = round(vol, 3)

        raw_ethanol = ws.cell(row=volume_rows["EtOH"], column=col).value
        try:
            ethanol = float(raw_ethanol)
        except (TypeError, ValueError):
            ethanol = 0.0

        ethanol_hv = int(ethanol)
        ethanol_lv = round(ethanol - ethanol_hv, 3)

        row_dict["EtOH_HV"] = ethanol_hv
        row_dict["EtOH_LV"] = ethanol_lv

        # Fill missing cols with 0
        for header in unique_headers:
            row_dict.setdefault(header, 0.0)

        # Add to master list
        rows.append(row_dict)

    # Create DataFrame
    df = pd.DataFrame(rows)
    
    csv_path = f'exp_templates/{RUN_NAME}_mantis_ready.csv'
    # Save to CSV
    df.to_csv(csv_path, index = False)
    print(f"Mantis-ready CSV saved as {csv_path}")


def open_excel_file(filepath):
    system = platform.system()
    if system == "Windows":
        os.startfile(filepath)
    elif system == "Darwin":  # macOS
        subprocess.call(["open", filepath])
    elif system == "Linux":
        subprocess.call(["xdg-open", filepath])
    else:
        print("Unsupported OS. Please open the file manually:", filepath)
